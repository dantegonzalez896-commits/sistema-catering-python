import sqlite3
from datetime import datetime, date, timedelta
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side

nombre_bd = "Catering.db"

with sqlite3.connect(nombre_bd) as conn:
    mi_cursor = conn.cursor()
    
    SQL_clientes = """
    CREATE TABLE IF NOT EXISTS clientes (
    id_cliente INTEGER PRIMARY KEY AUTOINCREMENT,
    nombres TEXT NOT NULL,
    apellidos TEXT NOT NULL
    );
    """
    
    SQL_platillos = """
    CREATE TABLE IF NOT EXISTS platillos (
    id_platillo INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre_platillo TEXT NOT NULL
    );
    """
    
    SQL_pedidos = """
    CREATE TABLE IF NOT EXISTS pedidos (
    folio INTEGER PRIMARY KEY AUTOINCREMENT,
    fecha_pedido TEXT NOT NULL,
    id_cliente INTEGER NOT NULL,
    turno TEXT NOT NULL,
    nombre_evento TEXT NOT NULL,
    estado INTEGER DEFAULT 1, 
    FOREIGN KEY (id_cliente) REFERENCES clientes (id_cliente)
    );
    """
    
    SQL_detalle_pedidos = """
    CREATE TABLE IF NOT EXISTS detalle_pedidos (
    id_detalle INTEGER PRIMARY KEY AUTOINCREMENT,
    folio INTEGER NOT NULL,
    id_platillo INTEGER NOT NULL,
    porciones INTEGER NOT NULL,
    FOREIGN KEY (folio) REFERENCES pedidos (folio),
    FOREIGN KEY (id_platillo) REFERENCES platillos (id_platillo)
    );
    """

    print("Todas las tablas del sistema Catering han sido validadas o creadas.")

def validar_inicio():
    """
    Valida la integridad física de las tablas al arrancar el programa.
    """
    sujetos = ["clientes", "platillos", "pedidos"]
    existentes = []

    with sqlite3.connect("Catering.db") as conn:
        cursor = conn.cursor()
        for tabla in sujetos:
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{tabla}'")
            if cursor.fetchone():
                existentes.append(tabla)

    if len(existentes) == 0:
        print("No se encontró almacenamiento previo. Se iniciará con un estado vacío")
        crear_tablas_iniciales()
    elif len(existentes) == len(sujetos):
        print("Estado anterior recuperado exitosamente")
    else:
        print("HA OCURRIDO UN ERROR DE ALMACENAMIENTO CRÍTICO")
        print("No se encuentran todos los sujetos del programa. Se cancelará la ejecución")
        import sys
        sys.exit()

def registrar_cliente():
    print("Registrar Nuevo Cliente")
    nombres = input("Ingrese los nombres: ").strip()
    apellidos = input("Ingrese los apellidos: ").strip()
    
    if not nombres or not apellidos:
        print("Error: Los datos no pueden quedar en blanco.")
        return   
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            instruccion_sql = "INSERT INTO clientes (nombres, apellidos) VALUES (?, ?)"  
            mi_cursor.execute(instruccion_sql, (nombres, apellidos))
            print("Registro exitoso.")
        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")

def registrar_platillo():
    print("Registrar Nuevo Platillo")
    nombre_platillo = input("Ingrese el nombre del platillo: ").strip()    
    
    if not nombre_platillo:
        print("Error: El nombre del platillo no puede estar en blanco.")
        return        
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()

        try:
            instruccion_sql = "INSERT INTO platillos (nombre_platillo) VALUES (?)"           
            mi_cursor.execute(instruccion_sql, (nombre_platillo,))            
            print("Platillo registrado exitosamente.")
        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")

def registrar_pedido():
    print("Registrar nuevo pedido")
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT COUNT(*) FROM clientes")
        
        if mi_cursor.fetchone()[0] == 0:
            print("Error: No hay clientes registrados en el sistema.")
            return
        mi_cursor.execute("SELECT COUNT(*) FROM platillos")

        if mi_cursor.fetchone()[0] == 0:
            print("Error: No hay platillos registrados en el sistema.")
            return
        print("Clientes Disponibles:")
        mi_cursor.execute("SELECT id_cliente, nombres, apellidos FROM clientes")
        for cliente in mi_cursor.fetchall():
            print(f"ID: {cliente[0]} | Nombre: {cliente[1]} {cliente[2]}")            
        try:
            cliente_id = int(input("Ingresa la Clave del cliente: "))
        except ValueError:
            print("Error: Debes ingresar un número entero.")
            return            
        mi_cursor.execute("SELECT 1 FROM clientes WHERE id_cliente = ?", (cliente_id,))
        
        if not mi_cursor.fetchone():
            print("Error: Clave de cliente no encontrada.")
            return
        fecha_str = input("Fecha del evento (MM/DD/YYYY): ")
        try:
            fecha_evento = datetime.strptime(fecha_str, "%m/%d/%Y").date()
            hoy = date.today()           
            if fecha_evento < (hoy + timedelta(days=2)):
                print("Error: El pedido debe realizarse con al menos 2 días de anticipación.")
                return
        except ValueError:
            print("Error: Formato de fecha no válido.")
            return
        turno = input("Escribe el turno (Mañana, Tarde, Noche): ").strip().capitalize()
        
        if turno not in ["Mañana", "Tarde", "Noche"]:
            print("Error: Turno inválido.")
            return
        mi_cursor.execute("""
            SELECT 1 FROM pedidos 
            WHERE id_cliente = ? AND fecha_pedido = ? AND turno = ? AND estado = 1
        """, (cliente_id, fecha_str, turno)) 
        
        if mi_cursor.fetchone():
            print("Error: El cliente ya tiene un evento programado para ese mismo turno y fecha.")
            return

        nombre_evento = input("Nombre del evento: ").strip()
        if not nombre_evento:
            print("Error: El evento debe tener un nombre.")
            return

        try:
            mi_cursor.execute("""
                INSERT INTO pedidos (fecha_pedido, id_cliente, turno, nombre_evento) 
                VALUES (?, ?, ?, ?)
            """, (fecha_str, cliente_id, turno, nombre_evento))
            
            folio_generado = mi_cursor.lastrowid
            
            agregar_comida = True 
            print("\n--- Agregar Platillos ---")

            while agregar_comida:
                print("Menú disponible:")
                mi_cursor.execute("SELECT id_platillo, nombre_platillo FROM platillos")
                for platillo in mi_cursor.fetchall():
                    print(f"Clave: {platillo[0]} - {platillo[1]}")
                         
                try:
                    id_platillo = int(input("Clave del platillo (0 para terminar): ")) 
                    if id_platillo == 0:
                        mi_cursor.execute("SELECT COUNT(*) FROM detalle_pedidos WHERE folio = ?", (folio_generado,))
                        if mi_cursor.fetchone()[0] > 0:
                            agregar_comida = False
                            break
                        else:
                            print("Error: Debes elegir al menos un platillo para el pedido.")
                            continue
                    
                    mi_cursor.execute("SELECT 1 FROM platillos WHERE id_platillo = ?", (id_platillo,))
                    if not mi_cursor.fetchone():
                        print("Error: El platillo seleccionado no existe.")
                        continue
                        
                    porciones = int(input("¿Cuántas porciones?: "))
                    if porciones <= 0:
                        print("Error: Las porciones deben ser mayores a 0.")
                        continue

                    mi_cursor.execute("SELECT porciones FROM detalle_pedidos WHERE folio = ? AND id_platillo = ?", (folio_generado, id_platillo))
                    registro_previo = mi_cursor.fetchone()

                    if registro_previo:
                        nuevas_porciones = registro_previo[0] + porciones
                        mi_cursor.execute("UPDATE detalle_pedidos SET porciones = ? WHERE folio = ? AND id_platillo = ?", (nuevas_porciones, folio_generado, id_platillo))
                        print(f"Platillo repetido: se sumaron las porciones. Total: {nuevas_porciones}")
                    else:
                        mi_cursor.execute("INSERT INTO detalle_pedidos (folio, id_platillo, porciones) VALUES (?, ?, ?)", (folio_generado, id_platillo, porciones))
                        print("Platillo agregado al pedido.")                
                except ValueError:
                    print("Error: Ingresa solo números enteros.")

            print(f"¡Pedido guardado exitosamente de forma directa en la base de datos! Folio asignado: {folio_generado}")
            
        except sqlite3.Error as e:
            print(f"Error crítico al guardar el pedido en la base de datos: {e}")

def reporte_pedidos_fechas():
    print("Reporte de Pedidos por Rango de Fechas")
    
    fecha_inicio_str = input("Ingrese fecha inicial (MM/DD/YYYY) o presione Enter para usar la fecha actual: ").strip()
    fecha_fin_str = input("Ingrese fecha final (MM/DD/YYYY) o presione Enter para usar la fecha actual: ").strip()
    hoy = date.today()
    
    try:
        if not fecha_inicio_str:
            fecha_inicio = hoy
        else:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%m/%d/%Y").date()
            
        if not fecha_fin_str:
            fecha_fin = hoy
        else:
            fecha_fin = datetime.strptime(fecha_fin_str, "%m/%d/%Y").date()
            
    except ValueError:
        print("Error: Formato de fecha no válido. Asegúrese de usar barras (MM/DD/YYYY).")
        return

    if fecha_inicio > fecha_fin:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    print(f"Generando reporte desde {fecha_inicio.strftime('%m/%d/%Y')} hasta {fecha_fin.strftime('%m/%d/%Y')}...\n")

    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            consulta = """
                SELECT p.folio, p.fecha_pedido, p.nombre_evento, c.nombres, c.apellidos, 
                       pl.nombre_platillo, d.porciones
                FROM pedidos p
                JOIN clientes c ON p.id_cliente = c.id_cliente
                JOIN detalle_pedidos d ON p.folio = d.folio
                JOIN platillos pl ON d.id_platillo = pl.id_platillo
                WHERE p.estado = 1
            """
            mi_cursor.execute(consulta)
            registros = mi_cursor.fetchall()
            
            datos_para_excel = []
            encontrados = False
            
            for r in registros:
                fecha_bd = datetime.strptime(r[1], "%m/%d/%Y").date()
                
                if fecha_inicio <= fecha_bd <= fecha_fin:
                    folio = r[0]
                    fecha = r[1]
                    evento = r[2]
                    cliente = f"{r[3]} {r[4]}"
                    platillo = r[5]
                    porciones = r[6]
                    
                    print(f"Folio: {folio} | Fecha: {fecha} | Cliente: {cliente} | Evento: {evento} | Platillo: {platillo} (Porciones: {porciones})")
                    
                    datos_para_excel.append([folio, fecha, cliente, evento, platillo, porciones])
                    encontrados = True
                    
            if not encontrados:
                print("No se encontraron pedidos en ese rango de fechas.")
                return 
            
            exportar = input("¿Desea exportar este reporte a MsExcel? (S/N): ").strip().upper()
            
            if exportar == "S":
                libro = openpyxl.Workbook()
                hoja = libro.active
                hoja.title = "Reporte Fechas"
                
                fuente_negrita = Font(bold=True)
                centrado = Alignment(horizontal="center", vertical="center")
                
                lado_doble = Side(border_style="double", color="000000")
                borde_doble = Border(top=lado_doble, bottom=lado_doble, left=lado_doble, right=lado_doble)
                
                encabezados = ["Folio", "Fecha", "Cliente", "Evento", "Platillo", "Porciones"]
                hoja.append(encabezados)
                
                for fila_datos in datos_para_excel:
                    hoja.append(fila_datos)
                
                total_filas = hoja.max_row 
                total_columnas = hoja.max_column
                
                for fila_idx in range(1, total_filas + 1):
                    for col_idx in range(1, total_columnas + 1):
                        celda_actual = hoja.cell(row=fila_idx, column=col_idx)
                        
                        celda_actual.border = borde_doble
                        
                        if fila_idx == 1:
                            celda_actual.font = fuente_negrita
                            celda_actual.alignment = centrado 
                            
                        elif col_idx in (1, 6):
                            celda_actual.alignment = centrado
                            
                hoja.column_dimensions["C"].width = 25
                hoja.column_dimensions["D"].width = 20
                hoja.column_dimensions["E"].width = 20
                
                nombre_archivo = f"Reporte_Pedidos_{hoy.strftime('%m%d%Y')}.xlsx"
                libro.save(nombre_archivo)
                print(f"¡Reporte exportado exitosamente como '{nombre_archivo}'!")
                
        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")

def reporte_clientes():
    print("\nREPORTE: CATÁLOGO DE CLIENTES")
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            mi_cursor.execute("SELECT id_cliente, nombres, apellidos FROM clientes ORDER BY apellidos, nombres")
            registros = mi_cursor.fetchall()
            
            if registros:
                print(f"{'ID':<6} | {'APELLIDOS':<20} | {'NOMBRES':<20}")
                print("-" * 50)
                for fila in registros:
                    print(f"{fila[0]:<6} | {fila[2]:<20} | {fila[1]:<20}")
            else:
                print("No hay clientes registrados en el sistema.") 
        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")

def reporte_platillos():
    print("\nREPORTE: CATÁLOGO DE PLATILLOS")
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            mi_cursor.execute("SELECT id_platillo, nombre_platillo FROM platillos ORDER BY nombre_platillo")
            registros = mi_cursor.fetchall()
            
            if registros:
                print(f"{'ID':<6} | {'PLATILLO':<30}")
                print("-" * 40)
                for fila in registros:
                    print(f"{fila[0]:<6} | {fila[1]:<30}")
            else:
                print("No hay platillos registrados en el sistema.")
        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")

def reporte_estadistico():
    print("\nREPORTE ESTADÍSTICO: PLATILLOS POR EVENTOS")
    
    fecha_ini = input("Fecha inicial (MM/DD/YYYY) [Enter para Hoy]: ").strip() or date.today().strftime("%m/%d/%Y")
    fecha_fin = input("Fecha final (MM/DD/YYYY) [Enter para Hoy]: ").strip() or date.today().strftime("%m/%d/%Y")

    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            mi_cursor.execute("""
                SELECT pl.nombre_platillo, COUNT(DISTINCT p.folio) as total_eventos
                FROM platillos pl
                JOIN detalle_pedidos d ON pl.id_platillo = d.id_platillo
                JOIN pedidos p ON d.folio = p.folio
                WHERE p.fecha_pedido BETWEEN ? AND ? AND p.estado = 1
                GROUP BY pl.id_platillo
                ORDER BY total_eventos DESC
            """, (fecha_ini, fecha_fin))
            
            registros = mi_cursor.fetchall()
            
            if registros:
                print(f"{'PLATILLO':<30} | {'CANT. EVENTOS':<15}")
                print("-" * 50)
                for fila in registros:
                    print(f"{fila[0]:<30} | {fila[1]:^15}")
            else:
                print("No hay datos para el rango seleccionado.") 
        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")

def inicializar_bd():
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            mi_cursor.execute(SQL_clientes)
            mi_cursor.execute(SQL_platillos)
            mi_cursor.execute(SQL_pedidos)
            mi_cursor.execute(SQL_detalle_pedidos)
        except sqlite3.Error as e:
            print(f"Error al crear las tablas: {e}")

def editar_evento():
    print("Editar Nombre del Evento")

    fecha_inicio_str = input("Ingrese fecha inicial (MM/DD/YYYY) o presione Enter para usar la fecha actual: ").strip()
    fecha_fin_str = input("Ingrese fecha final (MM/DD/YYYY) o presione Enter para usar la fecha actual: ").strip()

    hoy = date.today()

    try:
        if not fecha_inicio_str:
            fecha_inicio = hoy
        else:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%m/%d/%Y").date()
            
        if not fecha_fin_str:
            fecha_fin = hoy
        else:
            fecha_fin = datetime.strptime(fecha_fin_str, "%m/%d/%Y").date()
            
    except ValueError:
        print("Error: Formato de fecha no válido. Asegúrese de usar barras (MM/DD/YYYY).")
        return

    if fecha_inicio > fecha_fin:
        fecha_inicio, fecha_fin = fecha_fin, fecha_inicio

    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
            mi_cursor.execute("""
                SELECT p.folio, p.fecha_pedido, p.nombre_evento, c.nombres, c.apellidos
                FROM pedidos p
                JOIN clientes c ON p.id_cliente = c.id_cliente
            """)
            registros = mi_cursor.fetchall()

            folios_validos = []
            encontrados = False

            print("Pedidos encontrados en el rango de fechas:")
            for r in registros:
                fecha_bd = datetime.strptime(r[1], "%m/%d/%Y").date()

                if fecha_inicio <= fecha_bd <= fecha_fin:
                    print(f"Folio: {r[0]} | Fecha: {r[1]} | Evento Actual: {r[2]} | Cliente: {r[3]} {r[4]}")
                    folios_validos.append(r[0]) 
                    encontrados = True

            if not encontrados:
                print("No se encontraron pedidos en ese rango de fechas.")
                return

            try:
                folio_editar = int(input("Ingrese el folio del pedido a editar: "))
            except ValueError:
                print("Error: Se debe ingresar un número entero.")
                return

            if folio_editar not in folios_validos:
                print("Error: El folio ingresado no es válido o no pertenece al rango consultado.")
                return

            nuevo_nombre = input("Ingrese el nuevo nombre del evento: ").strip()
            if not nuevo_nombre:
                print("Error: El nombre del evento no puede quedar en blanco.")
                return

            mi_cursor.execute("""
                UPDATE pedidos
                SET nombre_evento = ?
                WHERE folio = ?
            """, (nuevo_nombre, folio_editar))

            print(f"¡El nombre del evento para el folio {folio_editar} se ha actualizado exitosamente!")

        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")


def cancelar_pedido():
    print("Cancelar un Pedido")
    fecha_str = input("Ingrese la fecha del pedido a cancelar (MM/DD/YYYY): ").strip()

    try:
        fecha_cancelar = datetime.strptime(fecha_str, "%m/%d/%Y").date()
        hoy = date.today()
        
        if fecha_cancelar < (hoy + timedelta(days=2)):
            print("Error: La cancelación exige al menos 2 días de anticipación a la fecha actual.")
            return
    except ValueError:
        print("Error: Formato de fecha no válido.")
        return

    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        try:
        
            mi_cursor.execute("""
                SELECT p.folio, c.nombres, c.apellidos, p.nombre_evento
                FROM pedidos p
                JOIN clientes c ON p.id_cliente = c.id_cliente
                WHERE p.fecha_pedido = ? AND p.estado = 1 
            """, (fecha_str,))
            
            registros = mi_cursor.fetchall()
            
            if not registros:
                print("No se encontraron pedidos activos para esa fecha.")
                return

            print("Listado de pedidos para esa fecha:")
            folios_validos = []
            for r in registros:
                print(f"Folio: {r[0]} | Cliente: {r[1]} {r[2]} | Evento/Turno: {r[3]}")
                folios_validos.append(r[0])

            opcion = input("\nIngrese el folio a cancelar (o escriba '0' para regresar al menú): ").strip()
            
            if opcion == "0":
                print("Operación abortada.")
                return
            
            try:
                folio_cancelar = int(opcion)
            except ValueError:
                print("Error: Se debe ingresar un número.")
                return

            if folio_cancelar not in folios_validos:
                print("Error: El folio elegido no corresponde a la lista.")
                return

            mi_cursor.execute("""
                SELECT p.folio, p.fecha_pedido, p.nombre_evento, c.nombres, c.apellidos
                FROM pedidos p
                JOIN clientes c ON p.id_cliente = c.id_cliente
                WHERE p.folio = ?
            """, (folio_cancelar,))
            detalle = mi_cursor.fetchone()
            
            print("DETALLE DEL PEDIDO")
            print(f"Folio: {detalle[0]}\nFecha: {detalle[1]}\nCliente: {detalle[3]} {detalle[4]}\nEvento: {detalle[2]}")
            
            confirmacion = input("¿Confirmar la cancelación definitiva? (S/N): ").strip().upper()
            
            if confirmacion == "S":
                mi_cursor.execute("UPDATE pedidos SET estado = 0 WHERE folio = ?", (folio_cancelar,))
                print(f"\n¡Pedido con folio {folio_cancelar} cancelado exitosamente del sistema!")
            else:
                print("Operación abortada.")

        except sqlite3.Error as e:
            print(f"Error de base de datos: {e}")



inicializar_bd()

def menu_reportes():
    "Submenú para la generación de reportes del sistema."
    while True:
        print("MENÚ DE REPORTES")
        print("1. Reporte de pedidos por rango de fechas")
        print("2. Reporte de clientes")
        print("3. Reporte de platillos")
        print("4. Reporte estadístico")
        print("5. Regresar al menú principal")
        opcion = input("Seleccione una opción: ").strip()
        
        if opcion == "1":
            reporte_pedidos_fechas()
        elif opcion == "2":
            reporte_clientes()
        elif opcion == "3":
            reporte_platillos()
        elif opcion == "4":
            reporte_estadistico()
        elif opcion == "5":
            print("Regresando al menú principal...")
            break 
        else:
            print("Opción inválida, intente de nuevo.")

def validar_salida():
    print("Salir del sistema")
    with sqlite3.connect("Catering.db") as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT COUNT(*) FROM clientes")
        c_clientes = mi_cursor.fetchone()[0]
        mi_cursor.execute("SELECT COUNT(*) FROM platillos")
        c_platillos = mi_cursor.fetchone()[0]
        mi_cursor.execute("SELECT COUNT(*) FROM pedidos")
        c_pedidos = mi_cursor.fetchone()[0]

        if c_clientes > 0 and c_platillos > 0 and c_pedidos > 0:
            confirmacion = input("¿Confirmar la salida permanente del sistema? (S/N): ").strip().upper()
            if confirmacion == "S":
                print("Almacenamiento permanente realizado. Saliendo del sistema...")
                return True
            else:
                print("Salida abortada. Regresando al menú.")
                return False
        else:
            print("Error: No se puede salir. Debe existir al menos un registro en clientes, platillos y pedidos.")
            return False

validar_inicio()

while True:
    print("SISTEMA DE CATERING")
    print("1. Registrar nuevo cliente")
    print("2. Registrar nuevo platillo")
    print("3. Registrar nuevo pedido")
    print("4. Reportes de pedidos")
    print("5. Editar nombre del evento")
    print("6. Cancelar evento")
    print("7. Salir")
    opcion = input("Ingrese una opción: ").strip()
    
    if opcion == "1":
        registrar_cliente()
    elif opcion == "2":
        registrar_platillo()
    elif opcion=="3":
        registrar_pedido()
    elif opcion == "4":
        menu_reportes()
    elif opcion == "5":
        editar_evento()
    elif opcion == "6":
        cancelar_pedido()
    elif opcion == "7":
        if validar_salida():
            break
    else:
        print("Opción invalida, intente de nuevo.")